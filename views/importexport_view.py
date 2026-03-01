"""
Import / Export View — Import da XLS (foglio DB_TNS) + Export su XLS.
"""
from __future__ import annotations

import io
import pandas as pd
import streamlit as st
import openpyxl

import db.connection as db

# Mapping colonne XLS → DB (foglio DB_TNS)
XLS_TO_DB_STRUTTURE = {
    "Codice":                    "codice",
    "DESCRIZIONE":               "descrizione",
    "CDCCOSTO":                  "cdc_costo",
    "UNITA' OPERATIVA PADRE ":   "codice_padre",
    "UNITA' OPERATIVA PADRE":    "codice_padre",
    "UNITA_OPERATIVA_PADRE":     "codice_padre",
    "Titolare":                  "titolare",
    "Livello":                   "livello",
    "Sede_TNS":                  "sede_tns",
    "Approvatore":               "approvatore",
    "Viaggiatore":               "viaggiatore",
    "Cassiere":                  "cassiere",
    "GruppoSind":                "gruppo_sind",
    "RuoliAFC":                  "ruoli_afc",
    "RuoliHR":                   "ruoli_hr",
}

XLS_TO_DB_DIPENDENTI = {
    "TxCodFiscale":              "codice_fiscale",
    "Titolare":                  "titolare",
    "Codice":                    "codice_nel_file",
    "Unità_Organizzativa":       "codice_struttura",
    "CDCCOSTO":                  "cdc_costo",
    "Livello":                   "livello",
    "Sede_TNS":                  "sede_tns",
    "Approvatore":               "approvatore",
    "Viaggiatore":               "viaggiatore",
    "Cassiere":                  "cassiere",
    "GruppoSind":                "gruppo_sind",
    "RuoliAFC":                  "ruoli_afc",
    "RuoliHR":                   "ruoli_hr",
}


def render() -> None:
    tab_import, tab_export = st.tabs(["📥 Import XLS", "📤 Export XLS"])

    with tab_import:
        _render_import()

    with tab_export:
        _render_export()


def _render_import() -> None:
    st.subheader("📥 Import da file XLS")
    st.info("Carica il file XLS originale (foglio **DB_TNS**). "
            "I record esistenti verranno aggiornati (upsert), i nuovi verranno creati.")

    uploaded = st.file_uploader("Scegli file XLS", type=["xls", "xlsx"], key="import_xls")
    if not uploaded:
        return

    try:
        # Leggi foglio DB_TNS
        if uploaded.name.endswith(".xlsx"):
            xf = pd.ExcelFile(uploaded, engine="openpyxl")
        else:
            xf = pd.ExcelFile(uploaded, engine="xlrd")

        if "DB_TNS" not in xf.sheet_names:
            st.error(f"Foglio 'DB_TNS' non trovato. Fogli disponibili: {xf.sheet_names}")
            return

        df = xf.parse("DB_TNS", dtype=str).fillna("")
        st.success(f"✅ Foglio DB_TNS letto — {len(df)} righe, {len(df.columns)} colonne")

        with st.expander("Anteprima prime 5 righe"):
            st.dataframe(df.head())

        # Separa strutture e dipendenti
        # Strutture: righe dove TxCodFiscale è vuoto
        # Dipendenti: righe dove TxCodFiscale è valorizzato
        has_cf = "TxCodFiscale" in df.columns
        if has_cf:
            df_str = df[df["TxCodFiscale"].str.strip() == ""].copy()
            df_dip = df[df["TxCodFiscale"].str.strip() != ""].copy()
        else:
            df_str = df.copy()
            df_dip = pd.DataFrame()

        st.write(f"📂 **{len(df_str)}** strutture  ·  👤 **{len(df_dip)}** dipendenti rilevati")

        col1, col2 = st.columns(2)
        import_str = col1.checkbox("Importa strutture", value=True)
        import_dip = col2.checkbox("Importa dipendenti", value=True)

        if st.button("🚀 Avvia import", type="primary"):
            progress = st.progress(0, text="Elaborazione...")

            total = (len(df_str) if import_str else 0) + (len(df_dip) if import_dip else 0)
            done = 0

            if import_str and not df_str.empty:
                rows_str = _map_df(df_str, XLS_TO_DB_STRUTTURE, pk="codice")
                # CDCCOSTO: detect numeric
                for r in rows_str:
                    try:
                        float(r.get("cdc_costo", ""))
                        r["cdc_costo_is_numeric"] = 1
                    except (ValueError, TypeError):
                        r["cdc_costo_is_numeric"] = 0
                n = db.upsert_strutture(rows_str)
                st.toast(f"✅ {n} strutture importate")
                done += len(rows_str)
                progress.progress(done / max(total, 1), text=f"Strutture: {n}")

            if import_dip and not df_dip.empty:
                rows_dip = _map_df(df_dip, XLS_TO_DB_DIPENDENTI, pk="codice_fiscale")
                for r in rows_dip:
                    try:
                        float(r.get("cdc_costo", ""))
                        r["cdc_costo_is_numeric"] = 1
                    except (ValueError, TypeError):
                        r["cdc_costo_is_numeric"] = 0
                n = db.upsert_dipendenti(rows_dip)
                st.toast(f"✅ {n} dipendenti importati")
                done += len(rows_dip)
                progress.progress(1.0, text=f"Dipendenti: {n}")

            st.success("🎉 Import completato!")
            st.rerun()

    except Exception as e:
        st.error(f"Errore durante l'import: {e}")


def _map_df(df: pd.DataFrame, mapping: dict[str, str], pk: str) -> list[dict]:
    """Rinomina le colonne secondo il mapping e rimuove righe senza PK."""
    # Trova il nome originale del PK
    pk_original = next((k for k, v in mapping.items() if v == pk), None)
    rows = []
    for _, row in df.iterrows():
        record: dict = {}
        for xls_col, db_col in mapping.items():
            if xls_col in df.columns:
                val = str(row.get(xls_col, "")).strip()
                record[db_col] = val if val not in ("", "nan", "None") else None
        if pk_original and not row.get(pk_original, "").strip():
            continue  # Salta righe senza PK
        # Rimuovi None per evitare problemi FK
        rows.append({k: v for k, v in record.items() if v is not None})
    return rows


def _render_export() -> None:
    st.subheader("📤 Export su XLS")
    st.info("Esporta strutture e dipendenti nel formato XLS compatibile con il file originale.")

    col1, col2 = st.columns(2)
    export_str = col1.checkbox("Esporta strutture", value=True)
    export_dip = col2.checkbox("Esporta dipendenti", value=True)

    if st.button("📥 Genera file XLS", type="primary"):
        try:
            wb = openpyxl.Workbook()
            wb.remove(wb.active)  # Rimuovi foglio vuoto default

            if export_str:
                strutture = db.fetch_strutture()
                df_str = pd.DataFrame(strutture)
                ws = wb.create_sheet("Strutture")
                if not df_str.empty:
                    ws.append(list(df_str.columns))
                    for _, row in df_str.iterrows():
                        ws.append([
                            float(v) if _is_numeric_float(str(v)) else str(v) if v is not None else ""
                            for v in row
                        ])

            if export_dip:
                dipendenti = db.fetch_dipendenti()
                df_dip = pd.DataFrame(dipendenti)
                ws2 = wb.create_sheet("Dipendenti")
                if not df_dip.empty:
                    ws2.append(list(df_dip.columns))
                    for _, row in df_dip.iterrows():
                        ws2.append([
                            float(v) if _is_numeric_float(str(v)) else str(v) if v is not None else ""
                            for v in row
                        ])

            # Salva in buffer
            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)

            st.download_button(
                label="⬇️ Scarica OrgPlus_export.xlsx",
                data=buf,
                file_name="OrgPlus_export.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        except Exception as e:
            st.error(f"Errore durante l'export: {e}")


def _is_numeric_float(val: str) -> bool:
    try:
        float(val)
        return "." in val or "e" in val.lower()
    except (ValueError, TypeError):
        return False
